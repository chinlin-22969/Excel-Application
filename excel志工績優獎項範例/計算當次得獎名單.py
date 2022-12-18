"""
File: Practice.py
Name: Who is eligible?
-----------------------
1. 抓取某範圍日期內所有志工時數(FILE1, 2)的總和（sum）。 return name: sum
2. 把總合丟入欲查詢的獎項(FILE3)function，僅搜集有符合獎項的志工名單。 return name: sum: rank
3. 把名單丟入function查詢是否曾有得獎紀錄(FILE3)，僅return過去沒有得獎過的志工名單，為最終得獎名單。 return name: sum: rank
4. 抓取得獎名單內的志工資料（name, birth, address, gender....）(FILE4)，在(FILE5)建立新的worksheet。 no return

"""

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# CONSTANTS
#   檔案表格格式設定
FILE1 = '服務時數總表.xlsx'
RS1 = 2  # row data start
CS1 = 3  # column data start
FILE2 = '年服務時數.xlsx'  # 須帶入欲查詢的年份
RS2 = 2  # row data start
CS2 = 3  # column data start
FILE3 = '得獎紀錄＆標準.xlsx'
FILE4 = '志工資料/志工資料.xlsx'
FILE5 = '歷年申請表單.xlsx'

#   當年度申請基本資訊（請參考政府來函）
GROUP_MEMBER = 10   # 正式志工人數
APPLY_YEAR = 111    # 申請年份（今年）
TROPHY = 'y獎'      # 欲申請獎項名稱(ex: x獎、y獎、z獎)

START_YEAR = 105
START_MONTH = 7
END_YEAR = 110
END_MONTH = 12


def main():
    name_sum = get_sum(START_YEAR, START_MONTH, END_YEAR, END_MONTH)
    eligible_d = get_ranks(TROPHY, name_sum)
    record_lst = trophy_record(TROPHY)
    print('時數達標準：', eligible_d)
    print('歷屆得獎：', record_lst)
    # 確認時數達標者，過去無此獎項紀錄
    final_d = get_final_d(record_lst, eligible_d)
    print('可申請者：', final_d)

    # 在歷年申請表單建立新的工作表，並貼上得獎者資訊
    load_data(final_d, APPLY_YEAR, TROPHY)


def get_sum(year1, month1, year2, month2):
    """
    Calculate each members' total service hours during the given range.

    :param year1: (int)start year
    :param month1: (int)start month
    :param year2: (int)end year
    :param month2: (int)end month

    :return: (dict) key -> value (name -> total service hour(year1/ month1 to year2/ month2))
    """
    # 若非整年度計算，則需另外拆開來，進行「月」為單位服務時數計算
    y1 = year1 + 1 if month1 > 1 else year1
    y2 = year2 - 1 if month2 < 12 else year2

    # 進行「年」為單位服務時數計算
    name_sum = get_sum_from_f2(y1, y2)

    # 進行「月」為單位服務時數計算
    # 起年
    if y1 != year1:
        wb = load_workbook(str(year1)+FILE2, data_only=True)
        ws = wb['總表']
        for row in range(RS2, GROUP_MEMBER + RS2):
            total = 0
            name = ws['B'+str(row)].value
            for column in range(month1+CS2-1, 12+CS2):  # month1～December
                char = get_column_letter(column)
                total += int(ws[char+str(row)].value)
            name_sum[name] += total
    # 迄年
    if y2 != year2:
        wb = load_workbook(str(year2)+FILE2, data_only=True)
        ws = wb['總表']
        for row in range(RS2, GROUP_MEMBER + RS2):
            total = 0
            name = ws['B'+str(row)].value
            for column in range(CS2, month2+CS2):  # January～month2
                char = get_column_letter(column)
                total += int(ws[char+str(row)].value)
            name_sum[name] += total
    return name_sum


def get_sum_from_f2(y1, y2):
    """
    :param y1:(int) 從年頭（一月份）開始算起的年份
    :param y2:(int) 計算到年尾（十二月份）的年份
    :return: (dict) key -> value (name -> total service hour(y1 to y2))
    """
    wb = load_workbook(FILE1, data_only=True)
    ws = wb.active

    # column 1)編號    2)姓名      3)100年  4).........
    # row    1) title 2)1st data 3)......
    d = {}
    for row in range(2, GROUP_MEMBER + 2):
        total = 0
        name = ws['B' + str(row)].value
        for column in range(y1 - 97, y2 - 97 + 1):  # 上限不包含故+1
            char = get_column_letter(column)
            total += int(ws[char + str(row)].value)
        d[name] = total
    return d


def get_ranks(trophy, name_sum):
    """
    Get the list of members whose service hours meet the criteria.
    loop over 'name_sum', if rank is not None, make it into 'name_sum_rank'.

    :param trophy:(str) name of the applying trophy
    :param name_sum:(dict) key -> value (name -> total service hour)

    :return name_sum_rank:(dict) key -> value (name -> [total service hour, '獎項名稱'])
    """

    wb = load_workbook(FILE3, data_only=True)
    ws = wb['獎項標準']
    # get column
    for column in range(1, 9):
        crit = get_column_letter(column)         # 各獎項標準
        if trophy == ws[crit+'1'].value:
            tro = get_column_letter(column - 1)  # 各獎項名稱

    # 'crit' infers the column of the criteria of the trophy
    # 'tro' infers the column of the trophy's name

            name_sum_rank = {}
            for name, total in name_sum.items():
                if ws[crit+'2'].value <= total:
                    rank = ws[tro+'2'].value
                elif ws[crit+'3'].value <= total < ws[crit+'2'].value:
                    rank = ws[tro+'3'].value
                elif ws[crit+'4'].value <= total < ws[crit+'3'].value:
                    rank = ws[tro+'4'].value
                else:
                    rank = None
                if rank is not None:
                    name_sum_rank[name] = [total, rank]
            return name_sum_rank
    return None


def trophy_record(trophy):
    """
    Make a list of the record of the trophy

    :param trophy: (str) name of the applying trophy

    :return: (lst) get the record of the trophy
    """
    wb = load_workbook(FILE3, data_only=True)
    ws = wb['歷屆得獎']
    history_lst = []

    # get_column
    for column in range(2, 7):
        char = get_column_letter(column)
        if trophy == ws[char+'1'].value:
            for row in range(2, 12):
                n = ws['B'+str(row)].value
                r = ws[char+str(row)].value
                history_lst.append((n, r))
    return history_lst


def get_final_d(record, d):
    """
    Check if the members have get this 獎項名稱 before, and return a final list of names.

    :param record: (lst)a list of the record of the trophy
    :param d: (dict)key -> value (name -> [total service hour, '獎項名稱'])

    :return: (dict)key -> value (name -> [total service hour, '獎項名稱'])
    """
    final_d = {}
    for ele in record:
        for name, sum_rank in d.items():
            if name == ele[0]:
                if not ele[1] or sum_rank[1] not in ele[1]:
                    final_d[name] = sum_rank
    return final_d


def load_data(d, year, trophy):
    """
    This function create a new worksheet (called ex: 109年度x獎申請表單) and copy the data of the final eligible member
    to the new worksheet. Also 該次獎項名稱 ＆ 服務時數。

    :param d:  (dict)key -> value (name -> [total service hour, '獎項名稱'])
    :param year: (int)the applying year
    :param trophy: (str)the applying trophy

    No returns
    """
    #  在歷年申請表單建立新的工作表
    wb1 = load_workbook(FILE5, data_only=True)  # create new excel file
    wb1.create_sheet(str(year)+'年度'+trophy+'申請表單')  # give name to the default ws
    ws1 = wb1[str(year)+'年度'+trophy+'申請表單']
    # 打開志工資料檔案
    wb2 = load_workbook(FILE4, data_only=True)
    ws2 = wb2['現任志工資料']
    # 在新工作表貼上title
    for row in ws2.iter_rows(min_row=1, max_row=1, min_col=1, max_col=10, values_only=True):
        ws1.append((*row, '服務時數', '獎項'))
    # 在新工作表上貼上得獎者資訊
    for name, sum_rank in d.items():
        for row in ws2.iter_rows(min_row=1, max_row=11, min_col=1, max_col=10, values_only=True):
            if row[1] == name:
                ws1.append((*row, sum_rank[0], sum_rank[1]))  # * is used to unpack tuple then add new ele
    # 確保格式相同 number_format
    for row in range(1, ws2.max_row+1):
        for col in range(1, ws2.max_column+1):
            c = ws2.cell(row=row, column=col)
            new_c = ws1.cell(row=row, column=col)
            new_c.number_format = c.number_format
    # 儲存
    wb1.save('歷年申請表單.xlsx')


def learning_pyxl():
    wb = load_workbook(FILE4)
    ws = wb.active
    print(ws.title)  # 工作表名稱
    print(ws.min_row, ws.max_row)  # row頭、row尾
    print(ws.min_column, ws.max_column)  # column頭、column尾
    print(ws.dimensions)  # 工作表範圍：左上格～右下格
    print(ws.cell(row=1, column=1).value)  # 印出此cell位置，value則印出此cell的值
    print(ws['A1'].value)  # 印出此cell位置，value則印出此cell的值

    title_index = {}
    index = 0
    # 可取得某一範圍的row資料，一個row一個tuple
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        for ele in row:
            title_index[ele] = index
            index += 1
    print(title_index)

    # 列印出所有data
    for row in ws.values:
        for value in row:
            print(value)


# DO NOT EDIT CODE BELOW THIS LINE #

if __name__ == "__main__":
    main()
